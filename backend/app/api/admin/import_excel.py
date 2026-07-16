"""
Admin API: Excel import for courses.

Routes:
    GET  /admin/import/template                       -- download a blank Excel template
    POST /admin/courses/{id}/imports/preview           -- upload & validate, return preview
    POST /admin/courses/{id}/imports/confirm           -- confirm a previewed import
    POST /admin/courses/{id}/import                    -- legacy one-shot import

Excel column mapping (A-J):
    A: chapter_name   (章节名称)
    B: open_at        (开放时间, optional)
    C: question_content (题干)
    D: option_a       (选项A)
    E: option_b       (选项B)
    F: option_c       (选项C)
    G: option_d       (选项D)
    H: option_e       (选项E, optional)
    I: correct_answer (正确答案, A-E)
    J: explanation    (解析, optional)
"""

from __future__ import annotations

import io
import logging
import time
import uuid
from datetime import datetime
from typing import Dict, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.db import get_db
from app.models import Chapter, Course, Question, QuestionImportBatch, User
from app.api.deps import get_admin_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/admin", tags=["Admin-Import"])


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXPECTED_COLUMNS = 10  # A through J
COLUMN_NAMES = [
    "章节名称",       # A
    "开放时间",       # B
    "题干",           # C
    "选项A",          # D
    "选项B",          # E
    "选项C",          # F
    "选项D",          # G
    "选项E",          # H (optional)
    "正确答案",       # I
    "解析",           # J (optional)
]

VALID_ANSWERS = {"A", "B", "C", "D", "E"}

# In-memory preview store: preview_id -> {data, expiry}
# Each entry auto-expires after 30 minutes.
_PREVIEW_STORE: Dict[str, dict] = {}
_PREVIEW_TTL = 30 * 60  # seconds


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------


class ImportStats(BaseModel):
    total_rows: int
    chapters_created: int
    questions_imported: int
    errors: List[str]


class PreviewError(BaseModel):
    row: int
    message: str


class ChapterPreview(BaseModel):
    name: str
    question_count: int


class PreviewResponse(BaseModel):
    preview_id: str
    total_rows: int
    valid_count: int
    error_count: int
    errors: List[PreviewError]
    chapters: List[ChapterPreview]


class ConfirmRequest(BaseModel):
    preview_id: str


class ConfirmResponse(BaseModel):
    batch_id: int
    imported_count: int
    chapter_count: int


# ---------------------------------------------------------------------------
# GET /admin/import/template
# ---------------------------------------------------------------------------


@router.get("/import/template")
async def download_template(
    admin: User = Depends(get_admin_user),
):
    """Download a blank Excel template for question import."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题目导入"

    # Write header row
    for col_idx, name in enumerate(COLUMN_NAMES, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = cell.font.copy(bold=True)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 20  # 章节名称
    ws.column_dimensions["B"].width = 18  # 开放时间
    ws.column_dimensions["C"].width = 50  # 题干
    ws.column_dimensions["D"].width = 30  # 选项A
    ws.column_dimensions["E"].width = 30  # 选项B
    ws.column_dimensions["F"].width = 30  # 选项C
    ws.column_dimensions["G"].width = 30  # 选项D
    ws.column_dimensions["H"].width = 30  # 选项E
    ws.column_dimensions["I"].width = 10  # 正确答案
    ws.column_dimensions["J"].width = 50  # 解析

    # Write one example row
    example = [
        "第一章 基础知识",          # 章节名称
        "2026-09-01 08:00",        # 开放时间
        "人体最大的器官是什么？",    # 题干
        "心脏",                    # 选项A
        "肝脏",                    # 选项B
        "皮肤",                    # 选项C
        "大脑",                    # 选项D
        "",                        # 选项E (optional)
        "C",                       # 正确答案
        "皮肤是人体面积最大的器官",  # 解析
    ]
    for col_idx, val in enumerate(example, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"},
    )


# ---------------------------------------------------------------------------
# POST /admin/courses/{course_id}/imports/preview
# ---------------------------------------------------------------------------


@router.post("/courses/{course_id}/imports/preview")
async def preview_import(
    course_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
) -> dict:
    """Upload an Excel file and return a validation preview.

    The file is parsed and validated but **nothing is written** to the
    database.  A ``preview_id`` is returned that must be passed to the
    confirm endpoint to actually persist the data.
    """
    # Verify course exists
    course = await db.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="课程不存在")

    # Read uploaded file
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传的文件为空")

    # Parse Excel
    rows = _parse_excel(content)
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 文件中没有数据行")

    # Validate and group
    chapters_data: Dict[str, List[dict]] = {}
    errors: List[dict] = []

    for row_idx, row in enumerate(rows, start=2):
        chapter_name = (row.get("chapter_name") or "").strip()
        open_at_raw = (row.get("open_at") or "").strip() or None
        question_content = (row.get("question_content") or "").strip()
        option_a = (row.get("option_a") or "").strip()
        option_b = (row.get("option_b") or "").strip()
        option_c = (row.get("option_c") or "").strip()
        option_d = (row.get("option_d") or "").strip()
        option_e = (row.get("option_e") or "").strip() or None
        correct_answer = (row.get("correct_answer") or "").strip().upper()
        explanation = (row.get("explanation") or "").strip() or None

        row_errors: List[str] = []

        # --- validation ---
        if not chapter_name:
            row_errors.append("章节名称为空")
        if not question_content:
            row_errors.append("题干为空")
        if not option_a:
            row_errors.append("选项A为空")
        if not option_b:
            row_errors.append("选项B为空")
        if not option_c:
            row_errors.append("选项C为空")
        if not option_d:
            row_errors.append("选项D为空")
        # option_e can be empty
        if correct_answer not in VALID_ANSWERS:
            row_errors.append(f"正确答案必须为 A-E，当前值: '{correct_answer}'")
        else:
            # Answer must correspond to a non-empty option
            option_map = {
                "A": option_a,
                "B": option_b,
                "C": option_c,
                "D": option_d,
                "E": option_e,
            }
            if not option_map.get(correct_answer):
                row_errors.append(f"正确答案为 {correct_answer}，但对应选项为空")

        if row_errors:
            for msg in row_errors:
                errors.append({"row": row_idx, "message": msg})
            continue

        # Parse open_at
        open_at = None
        if open_at_raw:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                try:
                    open_at = datetime.strptime(open_at_raw, fmt)
                    break
                except ValueError:
                    continue

        question_data = {
            "content": question_content,
            "option_a": option_a,
            "option_b": option_b,
            "option_c": option_c,
            "option_d": option_d,
            "option_e": option_e,
            "correct_answer": correct_answer,
            "explanation": explanation,
            "open_at": open_at,
        }

        chapters_data.setdefault(chapter_name, []).append(question_data)

    # Build chapter summaries
    chapter_summaries = [
        {"name": name, "question_count": len(questions)}
        for name, questions in chapters_data.items()
    ]

    valid_count = sum(len(qs) for qs in chapters_data.values())

    # Store preview data for the confirm step
    _cleanup_expired_previews()
    preview_id = uuid.uuid4().hex
    _PREVIEW_STORE[preview_id] = {
        "course_id": course_id,
        "filename": file.filename or "upload.xlsx",
        "admin_id": admin.id,
        "chapters_data": chapters_data,
        "total_rows": len(rows),
        "valid_count": valid_count,
        "error_count": len(errors),
        "created_at": time.time(),
    }

    return {
        "code": 200,
        "message": "预览完成",
        "data": {
            "preview_id": preview_id,
            "total_rows": len(rows),
            "valid_count": valid_count,
            "error_count": len(errors),
            "errors": errors,
            "chapters": chapter_summaries,
        },
    }


# ---------------------------------------------------------------------------
# POST /admin/courses/{course_id}/imports/confirm
# ---------------------------------------------------------------------------


@router.post("/courses/{course_id}/imports/confirm")
async def confirm_import(
    course_id: int,
    body: ConfirmRequest,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
) -> dict:
    """Confirm a previously previewed import.

    Retrieves the parsed data from the preview step and persists it
    inside a single database transaction:
      1. Create a ``QuestionImportBatch`` record.
      2. Deactivate old questions (set ``is_active=False``).
      3. Create chapters (if not already existing for this course).
      4. Create questions.
    """
    # Retrieve preview data
    preview = _PREVIEW_STORE.get(body.preview_id)
    if not preview:
        raise HTTPException(status_code=400, detail="预览已过期或不存在，请重新上传")

    if preview["course_id"] != course_id:
        raise HTTPException(status_code=400, detail="预览数据与课程不匹配")

    chapters_data: Dict[str, List[dict]] = preview["chapters_data"]
    if not chapters_data:
        raise HTTPException(status_code=400, detail="没有有效的题目数据")

    # Verify course still exists
    course = await db.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="课程不存在")

    batch = None
    try:
        # 1. Create import batch record
        batch = QuestionImportBatch(
            course_id=course_id,
            filename=preview["filename"],
            status="processing",
            question_count=preview["valid_count"],
            created_by=preview["admin_id"],
        )
        db.add(batch)
        await db.flush()
        await db.refresh(batch)

        # 2. Deactivate all existing questions for this course
        chapter_ids_result = await db.execute(
            select(Chapter.id).where(Chapter.course_id == course_id)
        )
        existing_chapter_ids = [row[0] for row in chapter_ids_result.fetchall()]

        if existing_chapter_ids:
            await db.execute(
                update(Question)
                .where(Question.chapter_id.in_(existing_chapter_ids))
                .values(is_active=False)
            )
            await db.flush()

        # 3. Create or reuse chapters, then create questions
        chapters_created = 0
        questions_imported = 0

        for chapter_sort, (chapter_name, questions) in enumerate(chapters_data.items()):
            # Find existing chapter with same name in this course
            result = await db.execute(
                select(Chapter).where(
                    Chapter.course_id == course_id,
                    Chapter.name == chapter_name,
                )
            )
            chapter = result.scalar_one_or_none()

            if chapter is None:
                # Use open_at from the first question that has it
                open_at = None
                for q in questions:
                    if q.get("open_at"):
                        open_at = q["open_at"]
                        break

                chapter = Chapter(
                    course_id=course_id,
                    name=chapter_name,
                    sort_order=chapter_sort,
                    open_at=open_at,
                    rag_enabled=False,
                )
                db.add(chapter)
                await db.flush()
                await db.refresh(chapter)
                chapters_created += 1

            # Create questions
            for q_idx, q_data in enumerate(questions):
                question = Question(
                    chapter_id=chapter.id,
                    import_batch_id=batch.id,
                    content=q_data["content"],
                    option_a=q_data["option_a"],
                    option_b=q_data["option_b"],
                    option_c=q_data["option_c"],
                    option_d=q_data["option_d"],
                    option_e=q_data["option_e"],
                    correct_answer=q_data["correct_answer"],
                    explanation=q_data["explanation"],
                    sort_order=q_idx,
                    is_active=True,
                )
                db.add(question)
                questions_imported += 1

        # Mark batch as done
        batch.status = "done"
        await db.flush()

        # Remove from preview store
        _PREVIEW_STORE.pop(body.preview_id, None)

        return {
            "code": 200,
            "message": "导入成功",
            "data": {
                "batch_id": batch.id,
                "imported_count": questions_imported,
                "chapter_count": chapters_created,
            },
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Excel import confirm failed for course %d", course_id)
        # Try to mark batch as failed
        try:
            if batch and batch.id:
                batch.status = "failed"
                batch.error_message = str(exc)[:500]
                await db.flush()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=f"导入失败: {exc}")


# ---------------------------------------------------------------------------
# POST /admin/courses/{id}/import  (legacy one-shot)
# ---------------------------------------------------------------------------


@router.post("/courses/{course_id}/import")
async def import_excel(
    course_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Import questions from an uploaded Excel file into a course.

    Behaviour:
    1. Parse the Excel file (columns A-J).
    2. Validate required columns and data.
    3. Group rows by chapter name.
    4. Within a single transaction:
       a. Deactivate all existing questions for the course.
       b. Create new chapters and questions from the spreadsheet.
    5. Return import statistics.
    """
    # Verify course exists
    course = await db.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="课程不存在")

    # Read uploaded file
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传的文件为空")

    # Parse Excel
    rows = _parse_excel(content)
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 文件中没有数据行")

    # Group by chapter name
    chapters_data: Dict[str, List[dict]] = {}
    errors: List[str] = []

    for row_idx, row in enumerate(rows, start=2):  # Excel row 2 = first data row
        chapter_name = (row.get("chapter_name") or "").strip()
        question_content = (row.get("question_content") or "").strip()
        option_a = (row.get("option_a") or "").strip()
        option_b = (row.get("option_b") or "").strip()
        option_c = (row.get("option_c") or "").strip()
        option_d = (row.get("option_d") or "").strip()
        option_e = (row.get("option_e") or "").strip() or None
        correct_answer = (row.get("correct_answer") or "").strip().upper()
        explanation = (row.get("explanation") or "").strip() or None

        # Validate required fields
        if not chapter_name:
            errors.append(f"第 {row_idx} 行: 章节名称为空")
            continue
        if not question_content:
            errors.append(f"第 {row_idx} 行: 题干为空")
            continue
        if not all([option_a, option_b, option_c, option_d]):
            errors.append(f"第 {row_idx} 行: 选项 A-D 不能为空")
            continue
        if correct_answer not in VALID_ANSWERS:
            errors.append(f"第 {row_idx} 行: 正确答案必须为 A/B/C/D/E，当前值: {correct_answer}")
            continue
        if correct_answer == "E" and not option_e:
            errors.append(f"第 {row_idx} 行: 正确答案为 E 但选项 E 为空")
            continue

        question_data = {
            "content": question_content,
            "option_a": option_a,
            "option_b": option_b,
            "option_c": option_c,
            "option_d": option_d,
            "option_e": option_e,
            "correct_answer": correct_answer,
            "explanation": explanation,
        }

        chapters_data.setdefault(chapter_name, []).append(question_data)

    if not chapters_data:
        raise HTTPException(
            status_code=400,
            detail="没有有效的题目数据" + (f"（{len(errors)} 行有错误）" if errors else ""),
        )

    # ── Transactional import ─────────────────────────────────────────
    try:
        # Create import batch
        batch = QuestionImportBatch(
            course_id=course_id,
            filename=file.filename or "upload.xlsx",
            status="processing",
            question_count=sum(len(qs) for qs in chapters_data.values()),
            created_by=admin.id,
        )
        db.add(batch)
        await db.flush()
        await db.refresh(batch)

        # Deactivate existing questions
        chapter_ids_result = await db.execute(
            select(Chapter.id).where(Chapter.course_id == course_id)
        )
        existing_chapter_ids = [row[0] for row in chapter_ids_result.fetchall()]
        if existing_chapter_ids:
            await db.execute(
                update(Question)
                .where(Question.chapter_id.in_(existing_chapter_ids))
                .values(is_active=False)
            )
            await db.flush()

        chapters_created = 0
        questions_imported = 0

        for chapter_sort, (chapter_name, questions) in enumerate(chapters_data.items()):
            # Find or create chapter
            result = await db.execute(
                select(Chapter).where(
                    Chapter.course_id == course_id,
                    Chapter.name == chapter_name,
                )
            )
            chapter = result.scalar_one_or_none()

            if chapter is None:
                chapter = Chapter(
                    course_id=course_id,
                    name=chapter_name,
                    sort_order=chapter_sort,
                    rag_enabled=False,
                )
                db.add(chapter)
                await db.flush()
                await db.refresh(chapter)
                chapters_created += 1

            # Create questions
            for q_idx, q_data in enumerate(questions):
                question = Question(
                    chapter_id=chapter.id,
                    import_batch_id=batch.id,
                    content=q_data["content"],
                    option_a=q_data["option_a"],
                    option_b=q_data["option_b"],
                    option_c=q_data["option_c"],
                    option_d=q_data["option_d"],
                    option_e=q_data["option_e"],
                    correct_answer=q_data["correct_answer"],
                    explanation=q_data["explanation"],
                    sort_order=q_idx,
                    is_active=True,
                )
                db.add(question)
                questions_imported += 1

        batch.status = "done"
        await db.flush()

        return {
            "code": 200,
            "message": "导入成功",
            "data": ImportStats(
                total_rows=len(rows),
                chapters_created=chapters_created,
                questions_imported=questions_imported,
                errors=errors,
            ),
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Excel import failed for course %d", course_id)
        raise HTTPException(status_code=500, detail=f"导入失败: {exc}")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _cleanup_expired_previews() -> None:
    """Remove expired entries from the preview store."""
    now = time.time()
    expired_keys = [
        k for k, v in _PREVIEW_STORE.items()
        if now - v.get("created_at", 0) > _PREVIEW_TTL
    ]
    for k in expired_keys:
        _PREVIEW_STORE.pop(k, None)


def _parse_excel(content: bytes) -> List[dict]:
    """Parse an Excel file and return a list of row dicts.

    自动检测列结构：
    - 标准格式 (A-J): 章节名称, 开放时间, 题干, 选项A-E, 正确答案, 解析
    - 护理题库格式 (A-L): 课程, 章节, 题型, 题干, 选项A-E, 正确答案, 解析, 所属章节
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # 读取表头判断格式
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(h or "").strip() for h in header]

    # 检测是否为护理题库格式（第1列是"课程"，第3列是"题型"）
    is_nursing_format = (
        len(header) >= 10 and
        ("课程" in header[0] or "course" in header[0].lower()) and
        ("题型" in header[2] or "type" in header[2].lower())
    )

    rows: List[dict] = []

    if is_nursing_format:
        # 护理题库格式: A=课程, B=章节, C=题型, D=题干, E-I=选项A-E, J=正确答案, K=解析
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if not row or all(cell is None for cell in row):
                continue

            row_dict = {
                "chapter_name": str(row[1] or "").strip() if len(row) > 1 else "",
                "open_at": None,  # 护理题库没有开放时间列
                "question_content": str(row[3] or "").strip() if len(row) > 3 else "",
                "option_a": str(row[4] or "").strip() if len(row) > 4 else "",
                "option_b": str(row[5] or "").strip() if len(row) > 5 else "",
                "option_c": str(row[6] or "").strip() if len(row) > 6 else "",
                "option_d": str(row[7] or "").strip() if len(row) > 7 else "",
                "option_e": str(row[8] or "").strip() if len(row) > 8 and row[8] else None,
                "correct_answer": str(row[9] or "").strip().upper() if len(row) > 9 else "",
                "explanation": str(row[10] or "").strip() if len(row) > 10 and row[10] else None,
            }
            rows.append(row_dict)
    else:
        # 标准格式: A=章节名称, B=开放时间, C=题干, D-H=选项A-E, I=正确答案, J=解析
        column_keys = [
            "chapter_name", "open_at", "question_content",
            "option_a", "option_b", "option_c", "option_d", "option_e",
            "correct_answer", "explanation",
        ]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=10, values_only=True)):
            if not row or all(cell is None for cell in row):
                continue

            row_dict = {}
            for col_idx, key in enumerate(column_keys):
                value = row[col_idx] if col_idx < len(row) else None
                if value is not None:
                    value = str(value)
                row_dict[key] = value
            rows.append(row_dict)

    wb.close()
    return rows
