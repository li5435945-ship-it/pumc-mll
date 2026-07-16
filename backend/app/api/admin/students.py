"""
Admin API: Student management.

Routes:
    GET    /admin/students              -- list students (paginated, searchable)
    POST   /admin/students              -- create a single student
    POST   /admin/students/batch        -- batch create students
    GET    /admin/students/export       -- export student list as CSV
    PUT    /admin/students/{id}         -- update student (nickname, password)
    DELETE /admin/students/{id}         -- delete student
"""

from __future__ import annotations

import csv
import io
import re
import logging
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession
from passlib.context import CryptContext

from app.db import get_db
from app.models import User
from app.api.deps import get_admin_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/admin", tags=["Admin-Students"])

pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$")


def _validate_email(email: str) -> bool:
    return bool(EMAIL_RE.match(email))


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------


class StudentOut(BaseModel):
    id: int
    email: str
    nickname: Optional[str] = None
    student_group: Optional[str] = None
    created_at: Optional[datetime] = None

    model_config = {"from_attributes": True}


class StudentCreate(BaseModel):
    email: str
    password: str
    nickname: Optional[str] = None
    student_group: Optional[str] = None


class StudentUpdate(BaseModel):
    nickname: Optional[str] = None
    password: Optional[str] = None


class StudentBatchCreate(BaseModel):
    students: List[StudentCreate]


class SkippedStudent(BaseModel):
    email: str
    reason: str


class StudentBatchResult(BaseModel):
    created: int
    skipped: List[SkippedStudent]


class PaginatedStudents(BaseModel):
    items: List[StudentOut]
    total: int
    page: int
    page_size: int


# ---------------------------------------------------------------------------
# GET /admin/students -- list all students (paginated, searchable)
# ---------------------------------------------------------------------------


@router.get("/students")
async def list_students(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    search: Optional[str] = Query(None, description="按邮箱或昵称搜索"),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Return a paginated list of students, optionally filtered by email/nickname."""
    base_query = select(User).where(User.role == "student")

    if search:
        pattern = f"%{search}%"
        base_query = base_query.where(
            or_(User.email.ilike(pattern), User.nickname.ilike(pattern))
        )

    # Total count
    count_query = select(func.count()).select_from(base_query.subquery())
    total = await db.scalar(count_query) or 0

    # Paginated results
    offset = (page - 1) * page_size
    result = await db.execute(
        base_query.order_by(User.id).offset(offset).limit(page_size)
    )
    students = result.scalars().all()

    return {
        "code": 200,
        "message": "success",
        "data": PaginatedStudents(
            items=[StudentOut.model_validate(s) for s in students],
            total=total,
            page=page,
            page_size=page_size,
        ),
    }


# ---------------------------------------------------------------------------
# POST /admin/students -- create a single student
# ---------------------------------------------------------------------------


@router.post("/students")
async def create_student(
    body: StudentCreate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Create a single student account."""
    # Validate email format
    if not _validate_email(body.email):
        raise HTTPException(status_code=400, detail="邮箱格式不正确")

    # Check duplicate
    existing = await db.scalar(select(User).where(User.email == body.email))
    if existing:
        raise HTTPException(status_code=409, detail=f"邮箱已存在: {body.email}")

    # Validate password
    if not body.password or len(body.password) < 6:
        raise HTTPException(status_code=400, detail="密码长度不能少于6位")

    # Create user
    user = User(
        email=body.email,
        password_hash=pwd_ctx.hash(body.password),
        nickname=body.nickname or body.email.split("@")[0],
        role="student",
        student_group=body.student_group,
    )
    db.add(user)
    await db.flush()
    await db.refresh(user)

    return {
        "code": 200,
        "message": "学生创建成功",
        "data": StudentOut.model_validate(user),
    }


# ---------------------------------------------------------------------------
# POST /admin/students/batch -- batch create students
# ---------------------------------------------------------------------------


@router.post("/students/batch")
async def batch_create_students(
    body: StudentBatchCreate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Batch create student accounts. Validates all entries, skips duplicates."""
    if not body.students:
        raise HTTPException(status_code=400, detail="学生列表不能为空")

    created_count = 0
    skipped: List[SkippedStudent] = []

    # Pre-fetch existing emails for this batch to avoid N+1 queries
    batch_emails = [s.email for s in body.students]
    result = await db.execute(
        select(User.email).where(User.email.in_(batch_emails))
    )
    existing_emails = {row[0] for row in result.all()}

    for entry in body.students:
        # Validate email format
        if not _validate_email(entry.email):
            skipped.append(SkippedStudent(email=entry.email, reason="邮箱格式不正确"))
            continue

        # Check duplicate (from DB or already-created in this batch)
        if entry.email in existing_emails:
            skipped.append(SkippedStudent(email=entry.email, reason="邮箱已存在"))
            continue

        # Validate password
        if not entry.password or len(entry.password) < 6:
            skipped.append(SkippedStudent(email=entry.email, reason="密码长度不能少于6位"))
            continue

        # Create user
        user = User(
            email=entry.email,
            password_hash=pwd_ctx.hash(entry.password),
            nickname=entry.nickname or entry.email.split("@")[0],
            role="student",
        )
        db.add(user)
        existing_emails.add(entry.email)  # prevent duplicate within same batch
        created_count += 1

    await db.flush()

    return {
        "code": 200,
        "message": f"批量创建完成: 成功 {created_count}，跳过 {len(skipped)}",
        "data": StudentBatchResult(
            created=created_count,
            skipped=skipped,
        ),
    }


# ---------------------------------------------------------------------------
# GET /admin/students/export -- export student list as CSV
# ---------------------------------------------------------------------------


@router.get("/students/export")
async def export_students(
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Export all students as a CSV file."""
    result = await db.execute(
        select(User)
        .where(User.role == "student")
        .order_by(User.id)
    )
    students = result.scalars().all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["ID", "邮箱", "昵称", "注册时间"])

    for s in students:
        writer.writerow([
            s.id,
            s.email,
            s.nickname or "",
            s.created_at.strftime("%Y-%m-%d %H:%M:%S") if s.created_at else "",
        ])

    buf.seek(0)

    # Wrap StringIO in BytesIO for StreamingResponse
    byte_buf = io.BytesIO(buf.getvalue().encode("utf-8-sig"))

    return StreamingResponse(
        byte_buf,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=students.csv"},
    )


# ---------------------------------------------------------------------------
# PUT /admin/students/{id} -- update student
# ---------------------------------------------------------------------------


@router.put("/students/{student_id}")
async def update_student(
    student_id: int,
    body: StudentUpdate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Update a student's nickname and/or password."""
    user = await db.get(User, student_id)
    if not user:
        raise HTTPException(status_code=404, detail="学生不存在")

    if user.role != "student":
        raise HTTPException(status_code=400, detail="该用户不是学生角色")

    update_data = body.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="没有需要更新的字段")

    if "nickname" in update_data:
        user.nickname = update_data["nickname"]

    if "password" in update_data:
        password = update_data["password"]
        if not password or len(password) < 6:
            raise HTTPException(status_code=400, detail="密码长度不能少于6位")
        user.password_hash = pwd_ctx.hash(password)

    await db.flush()
    await db.refresh(user)

    return {
        "code": 200,
        "message": "学生信息更新成功",
        "data": StudentOut.model_validate(user),
    }


# ---------------------------------------------------------------------------
# DELETE /admin/students/{id} -- delete student
# ---------------------------------------------------------------------------


@router.delete("/students/{student_id}")
async def delete_student(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Delete a student account."""
    user = await db.get(User, student_id)
    if not user:
        raise HTTPException(status_code=404, detail="学生不存在")

    if user.role != "student":
        raise HTTPException(status_code=400, detail="不能删除非学生用户")

    await db.delete(user)
    await db.flush()

    return {
        "code": 200,
        "message": "学生已删除",
        "data": {"student_id": student_id},
    }


# ---------------------------------------------------------------------------
# POST /admin/students/import -- import students from Excel
# ---------------------------------------------------------------------------


@router.get("/students/import-template")
async def download_student_import_template(
    admin: User = Depends(get_admin_user),
):
    """Download a blank Excel template for student import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生导入"

    # Header style
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header row with style
    headers = ["邮箱", "密码", "分组", "姓名"]
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 15

    # Write example row with style
    example = ["student@example.com", "123456", "A", "张三"]
    for col_idx, val in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_import_template.xlsx"},
    )


@router.post("/students/import")
async def import_students_from_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_admin_user),
):
    """Import students from an Excel file.

    Supported formats:
        3 columns: A=email, B=password, C=nickname
        4 columns: A=email, B=password, C=group, D=nickname
    """
    import openpyxl
    from fastapi import File, UploadFile

    # Validate file extension
    filename = file.filename or "upload.xlsx"
    if not filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件 (.xlsx 或 .xls)")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传的文件为空")

    # Parse Excel
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法解析 Excel 文件: {e}")

    # Detect format: check if header row has 3 or 4 columns
    header = [str(cell.value or "").strip() for cell in ws[1]]
    max_col = len(header)

    # Read rows (skip header)
    rows = list(ws.iter_rows(min_row=2, max_col=max_col, values_only=True))
    wb.close()

    if not rows:
        raise HTTPException(status_code=400, detail="Excel 中没有数据行")

    # Process students
    created_count = 0
    skipped: List[dict] = []

    # Pre-fetch existing emails
    all_emails = [str(row[0] or "").strip() for row in rows if row[0]]
    if all_emails:
        result = await db.execute(
            select(User.email).where(User.email.in_(all_emails))
        )
        existing_emails = {row[0] for row in result.all()}
    else:
        existing_emails = set()

    for row_idx, row in enumerate(rows, start=2):
        email = str(row[0] or "").strip()
        password = str(row[1] or "").strip()
        # Support both 3-column (email, password, nickname) and 4-column (email, password, group, nickname) formats
        nickname = None
        student_group = None
        if max_col >= 4:
            # 4-column format: email, password, group, nickname
            student_group = str(row[2] or "").strip() if len(row) > 2 and row[2] else None
            nickname = str(row[3] or "").strip() if len(row) > 3 and row[3] else None
        else:
            # 3-column format: email, password, nickname
            nickname = str(row[2] or "").strip() if len(row) > 2 and row[2] else None

        # Validate email
        if not email:
            skipped.append({"row": row_idx, "email": "", "reason": "邮箱为空"})
            continue

        if not _validate_email(email):
            skipped.append({"row": row_idx, "email": email, "reason": "邮箱格式不正确"})
            continue

        # Check duplicate
        if email in existing_emails:
            skipped.append({"row": row_idx, "email": email, "reason": "邮箱已存在"})
            continue

        # Validate password
        if not password or len(password) < 6:
            skipped.append({"row": row_idx, "email": email, "reason": "密码长度不能少于6位"})
            continue

        # Create user
        user = User(
            email=email,
            password_hash=pwd_ctx.hash(password),
            nickname=nickname or email.split("@")[0],
            role="student",
            student_group=student_group,
        )
        db.add(user)
        existing_emails.add(email)
        created_count += 1

    await db.flush()

    return {
        "code": 200,
        "message": f"导入完成: 成功 {created_count}，跳过 {len(skipped)}",
        "data": {
            "created": created_count,
            "skipped": skipped,
        },
    }


# ---------------------------------------------------------------------------
# Session management
# ---------------------------------------------------------------------------


@router.get("/sessions")
async def list_online_sessions(
    admin: User = Depends(get_admin_user),
):
    """List all online user sessions from Redis."""
    from app.redis import get_redis

    redis = get_redis()
    if not redis:
        return {
            "code": 200,
            "message": "Redis 不可用",
            "data": {
                "sessions": [],
                "count": 0,
            },
        }

    sessions = []

    # Scan for session:* keys
    async for key in redis.scan_iter("session:*"):
        user_id = key.split(":")[1]
        sessions.append({
            "user_id": int(user_id),
        })

    return {
        "code": 200,
        "message": "ok",
        "data": {
            "sessions": sessions,
            "count": len(sessions),
        },
    }


@router.delete("/sessions/{user_id}")
async def kick_user_session(
    user_id: int,
    admin: User = Depends(get_admin_user),
):
    """Kick a user by deleting their Redis session."""
    from app.redis import get_redis

    redis = get_redis()
    if not redis:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Redis 不可用，无法踢人",
        )

    session_key = f"session:{user_id}"
    deleted = await redis.delete(session_key)

    if deleted:
        return {
            "code": 200,
            "message": f"用户 {user_id} 已被踢下线",
            "data": {"user_id": user_id, "kicked": True},
        }
    else:
        return {
            "code": 200,
            "message": f"用户 {user_id} 没有活跃会话",
            "data": {"user_id": user_id, "kicked": False},
        }
